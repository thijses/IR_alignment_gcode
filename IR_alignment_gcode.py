"""
quick manual:
- enter the COM ports it asks for in the commandline, then tab to the GUI window
- press 'h' to auto-home the printer (NOTE: window will not respond untill it's done)
- press 'r' to reset position (and write position)
- if offset is not calibrated, use 'WASD'+'q','e' keys to move the head untill the LED and photodiode look aligned (doesn't need to be 100% perfect)
- press spacebar to start testing (can be paused with spacebar as well)
- once it's done it will show a plot (matplotlib) and save to excel
- if it crashses, it will attempt to save the data it gathered so far to 'output.xlsx'
- you can pause the testing and press 'v' to show a graph at any time
- press 'l' to disable steppers (if you're done, just to avoid overheating)
- press 'k' to manually save to an excel file (can be done mid-test, but you should want to pause)
- you can move the view with middle-mouse-button dragging, zoom by scrolling, turn off the grid with 'g' and change zoom mode (centered vs mouse-bound) with 'z'

required libraries:
- numpy (duh) @ 1.24.2
- pyserial @ 3.5
- openpyxl @ 3.1.2 (for saving to excel files)
- opencv-python @ 4.7.0.72 (for visualization)
- matplotlib @ 3.7.1 (for graphs)


"""

## TODO:
# - threading (only useful for the UI stuff)


import numpy as np
# from typing import Callable # just for type-hints to provide some nice syntax colering
import time
import datetime # just for generating output filenames automatically

import serial # for printer commands AND for testing IR
import serial.tools.list_ports # just for listing the available COM ports (debug)

import gcode_struff as GC # (my own code) placed in a seperate file for legibility

SERIAL_TIMEOUT_DEFAULT = 0.010 # 10ms default serial timeout is a little low, but it makes no-response results faster to determine. (NOTE: baud rate assurance added later)
## NOTE: IMPORANT: changing any serial.Serial class parameters (such as baudrate or timeout) may result in some garbage data being transmitted (specifically on Arduinos using an Atmega16u2 as UART bridge!)

def initSerial(COMport:str, baud:int, timeout:float=SERIAL_TIMEOUT_DEFAULT) -> serial.Serial | None:
    """ attempt to connect to a serial port with a given name """
    try:
        serialObj = serial.Serial()
        serialObj.baudrate = baud
        serialObj.timeout = timeout # is used whenever you use serialObj.read() (at least if you're not saturated by the number of bytes .in_waiting)
        serialObj.rts = 0
        serialObj.dtr = 0 
        serialObj.port = COMport
        print("connecting to serial port:", serialObj.port)
        serialObj.open() # try to open the serial
        return(serialObj if serialObj.is_open else None)
    except Exception as excep:
        print("failed to connect to to COMport:", COMport,"   exception:", excep)
        return(None)

## other printer-related functions:
def waitForOK(serialObj:serial.Serial, timeout:float=0.5) -> tuple[bool,bytes]:
    """ wait for GCODE_MARLIN_OK to be received \n
        returns: (whether it was received  ,and,  what data it reveived (should end with GCODE_MARLIN_OK) )"""
    startTime = time.time(); atLeastOnce=False;  readData = b''
    while((((time.time() - startTime) < timeout) or (atLeastOnce == False)) and (not readData.endswith(GC.GCODE_MARLIN_OK))): # read untill timeout or OK found
        atLeastOnce = True # make sure this loop runs AT LEAST once, regardless of stange time.time() behavior or low timout values
        readData += serialObj.read_until(GC.GCODE_MARLIN_OK)
    return(readData.endswith(GC.GCODE_MARLIN_OK), readData) # TODO: check data!
def autoHome(serialObj:serial.Serial, autohomeTimeout:float=30.0) -> bool:
    """ attempt to auto-home """
    serialObj.write(GC.GCODE_AUTO_HOME); time.sleep(0.025) # NOTE: sleep() needed becuase pyserial can interrupt its own write cycles with _reconfigure_port stuff!
    print("auto-homing... (wait for printer to finish)")
    success, readData = waitForOK(serialObj, autohomeTimeout)
    if(not readData.endswith(GC.GCODE_MARLIN_OK)):     print("autoHome waitForOK() returned:", readData)
    return(success)
def disableSteppers(serialObj:serial.Serial, timeout:float=0.25) -> bool:
    """ disable the steppers (for panic/debugging purposes) """
    serialObj.write(GC.GCODE_DISABLE_STEPPERS); time.sleep(0.025) # NOTE: sleep() needed becuase pyserial can interrupt its own write cycles with _reconfigure_port stuff!
    success, readData = waitForOK(serialObj, timeout)
    if(not readData.endswith(GC.GCODE_MARLIN_OK)):     print("disableSteppers waitForOK() returned:", readData)
    return(success)
def disableAutoReports(serialObj:serial.Serial, timeout:float=0.25) -> bool:
    """ set auto-reporting of position and temperature to disabled, as this code requests those manually """
    success = True;
    for gcode in [GC.GCODE_DISABLE_AUTO_REPORT_POSITION, GC.GCODE_DISABLE_AUTO_REPORT_TEMPERATURE]:
        serialObj.write(gcode); time.sleep(0.025) # NOTE: sleep() needed becuase pyserial can interrupt its own write cycles with _reconfigure_port stuff!
        success, readData = waitForOK(serialObj, timeout)
        if(not readData.endswith(GC.GCODE_MARLIN_OK)):     print("disableAutoReports waitForOK() returned:", readData)
    return(success)
def getCurrentPosition(serialObj:serial.Serial, timeout:float=0.5) -> tuple[bool,tuple[float,float,float],tuple[float,float,float]]:
    serialObj.write(GC.GCODE_GET_CURRENT_POSITION)
    ## it will send the response, followed by GCODE_MARLIN_OK (seperate lines). You could read them seperately:
    # startTime = time.time(); atLeastOnce=False;  readData = b''
    # while((((time.time() - startTime) < timeout) or (atLeastOnce == False)) and (not readData.endswith(b'\n'))):
    #     atLeastOnce = True # make sure this loop runs AT LEAST once, regardless of stange time.time() behavior or low timout values
    #     readData += serialObj.readline() # read one line (wait untill b'\n')
    ## instead, i will just read untill the GCODE_MARLIN_OK and then strip that part off
    success, readData = waitForOK(serialObj, timeout) # it will also reply with an ok
    if(not success):    print("getCurrentPosition() failed");  return(False, (0,0,0), (0,0,0))
    return(GC.parseM114(readData.strip(GC.GCODE_MARLIN_OK).split(GC.GCODE_MARLIN_OK)[0])) # attempt to parse the data and return the results

## IR testing functions:
def IRresponseTest(IR_TX_serial:serial.Serial, IR_RX_serial:serial.Serial|None=None) -> float:
    """ test IR communication """
    if(IR_RX_serial is None):
        IR_RX_serial = IR_TX_serial
    goodCounter:int = 0
    IR_RX_serial.flush() # library flush
    while(IR_RX_serial.in_waiting > 0): IR_RX_serial.read(IR_RX_serial.in_waiting); time.sleep(IR_RX_serial.timeout) # manual flush
    for i in range(256):
        data:bytes = i.to_bytes(1,'big')
        IR_TX_serial.write(data)
        readData:bytes = IR_RX_serial.read(1)
        if(readData == data):
            goodCounter += 1
        elif((abs(data[0] - readData[0]) < 2) if (len(readData) >= 1) else False):   print("IRresponseTest bad data, but (numerically) close!:", readData, "!=", data);  time.sleep(IR_RX_serial.timeout)
        # else:   print("IRresponseTest bad data:", readData, "!=", data)
        while(IR_RX_serial.in_waiting > 0): print("discarding excess data:", IR_RX_serial.read(IR_RX_serial.in_waiting), "(after looking for", data, ")");  time.sleep(IR_RX_serial.timeout)
        try: # don't want my excessive debugging effors to crash things
            if(readData != data):   badDataPattern[i] += 1
        except Exception as excep: doNothing=0; print("badDataPattern writing went wrong:", excep)
    return(goodCounter / 256)
# def testIR(IR_TX_serial:serial.Serial, IR_RX_serial:serial.Serial|None=None) -> float: # deconstructed into drawing loop (for now)

## matplotlib visualization:
import matplotlib.pyplot as plt # for displaying results
def plot5D(list5D: dict[int,list[tuple[float,float,float,float]]]):
    """ plot multiple subplots of 3D grid-scatter graphs \n
        list5D is a dict with keys=baud_rate and values= list like [[x,y,z,data], etc.] """
    import matplotlib.pyplot as plt # for displaying results
    if(len(list5D[tuple(list5D.keys())[0]]) < 1):
        print("can't plot empty list");  return
    def _plot4D(subplotAx:plt.Axes, list4D: list[tuple[float,float,float,float]]): # with inspiration from online examples
        """ plot values in a 3D grid-scatter \n
            list4D is a 2D list, formatted like [[x,y,z,data], etc.] """
        ## scatter plot:
        list4Dnp = np.array(list4D)
        x = list4Dnp[:, 0];  y = list4Dnp[:, 1];  z = list4Dnp[:, 2];  c = list4Dnp[:, 3];  s = (100 * c*IRtestHorizontalStepsize) # this format is plottable with a 3D scatter plot
        return(subplotAx.scatter(x, y, z, c=c, s=s, cmap='RdYlGn')) # plot! (also returns colorBar(?))
        # ## flattened contour plots: TODO (fix, i was almost there)!!!
        # dataZsplit:dict[float,list[tuple[float,float,float]]] = {} # split data based on Z height to discrete levels (NOTE: involves a bit of nasty float comparison)
        # for x,y,z,mes in list4D:
        #     if(z in dataZsplit): dataZsplit[z].append((x,y,mes))
        #     else:   dataZsplit[z] = [(x,y,mes)]
        # for z in dataZsplit:
        #     npArr = np.array(dataZsplit[z])
        #     npArr = npArr.round(2)
        #     X, Y, C = np.meshgrid(npArr[:,0], npArr[:,1], npArr[:,2]) # seems excessive to me, but matplotlib likes meshgrid...
        #     ax.contourf(X[:,:,0], Y[:,:,0], C[:,:,0], offset=z)
    figRows:int = max(int(len(list5D) ** 0.5), 1) # square root, rounded down
    figCols:int = max(len(list5D) // figRows, 1)
    # print("rows, cols:", figRows, figCols)
    fig = plt.figure(figsize=(5, 5))
    fig.suptitle("IR UART transfer success versus x,y,z offset")
    for i in range(len(list5D)): # for every key
        ax = fig.add_subplot(figRows, figCols, i+1, projection="3d")
        ax.set_title(tuple(list5D.keys())[i]);  ax.set_xlabel("x misalignment [mm]");  ax.set_ylabel("y misalignment [mm]");  ax.set_zlabel("z distance [mm]")
        img = _plot4D(ax, list5D[tuple(list5D.keys())[i]])
        # fig.colorbar(img) # TODO: individual color bars? constant color map (0~1)?
    plt.tight_layout()
    plt.show()

## excel stuff
def saveToExcel(list5D: dict[int,list[tuple[float,float,float,float]]], filename="output.xlsx"):
    """ save recorded data to a fancy excel file """
    import openpyxl as opxl
    def makeSheet(Wsheet, list4D: list[tuple[float,float,float,float]]):
        Wsheet.append(['x [mm]', 'y [mm]', 'z [mm]', 'measurement (0~1)']) # write column titles
        for i in range(len(list4D)):
            Wsheet.append(list4D[i])
    Wbook = opxl.Workbook() # create excel storage object
    while(len(Wbook.sheetnames) > 0): Wbook.remove(Wbook[Wbook.sheetnames[0]]) # delete the default 'Sheet'
    for key in list5D:
        Wsheet = Wbook.create_sheet(str(key))
        makeSheet(Wsheet, list5D[key])
    try:
        Wsheet = Wbook.create_sheet("badDataPattern");  Wsheet.append(['data byte', 'bad data counter'])
        for i in range(len(badDataPattern)):  Wsheet.append([i, badDataPattern[i]])
    except Exception as excep: doNothing=0; print("couldn't save badDataPattern to excel", excep)
    Wbook.save(filename if filename.endswith(".xlsx") else (filename+".xlsx"))
def generateFileName(list5D: dict[int,list[tuple[float,float,float,float]]]) -> str:
    """ optional function for auto-generating filenames (to keep track of data)"""
    filename = datetime.datetime.now().strftime("%Y-%m-%d_%H;%M;%S_")
    filename += str(tuple(list5D.keys())[0]) # alwyas include the first baud rate
    if(len(list5D) > 1): # only if multiple baud rates were actually used
        filename += "-" + str(tuple(list5D.keys())[-1])
    return(filename + ".xlsx")
def readFromExcel(filename:str):
    """ read data from an excel file (for extra post-measurement graph-making purposes) """
    import openpyxl as opxl
    Wbook = opxl.open(filename,read_only=True)
    list5D: dict[int,list[tuple[float,float,float,float]]] = {} # {baud : [(x,y,z,data), etc.]}
    for Wsheet in Wbook:
        if(str(Wsheet.title).find("badDataPattern") >= 0):
            continue # skip importing this sheet (it doesn't go in list5D anyway)
        # print("importing sheet:", Wsheet.title)
        try:
            baud = int(Wsheet.title)
            list5D[baud] = list(Wsheet.values)[1:] # put excel data in (newly created) dict key (skipping the first line, which holds the column names)
        except Exception as excep:
            print("FAILED to import sheet:", Wsheet.title, "  exception:", excep)
    Wbook.close()
    return(list5D)




if __name__ == "__main__": # an example of how this file may be used
    try:

        # baudRatesToTest:tuple[int] = (4800, 9600, 19200, 38400, 57600, 115200) # this is an important user setting to set.
        baudRatesToTest:tuple[int] = (9600, ) # the code below supports dynamic baud rate changes. However, the PCB being tested currently does not (you have to flash new FW to change baud rate)

        IRtestHorizontalStepsize = 0.5 # (mm) horizontal (x,y) movement step between measurements
        IRtestVerticalStepsize   = 0.5 # (mm) vertical (z) movement step (upwards) between measurements
        IRtestVerticalDistMax    = 10.0 # (mm) maximum vertical (z) distance to reach during test (vertical movement is only upwards, for obvious reasons)
        IRtestHorizontalDistMax  = 10.0 # (mm) maximum horizontal (x,y) offset in in both directions to move during test (set extra wide, as practical limits are dynamically calculated)
        IRtestContinuationThresh = 127/256 # it will keep spiraling until no meausrements in the past rotation are above this value
        IRtestVertStopThresh = IRtestVerticalStepsize * 10 # (mm) if absolutely 0 datapoints are above IRtestContinuationThresh for serveral Z steps (this), stop the test early
        IRtestPasses = 1 # how many times to repeat the test (results are simply averaged (for now)). 1 should be fine

        printerBaud:int = 250000 # semi-modern Marlin printers use 250000, older ones might use 115200, really modern ones might go above 250000
        printerSafeFeedrate:float = 1200 # (mm/min) feedrate at which things are unlikely to break
        printerJogFeedrate:float = printerSafeFeedrate # (mm/min) feedrate for movements requested through keyboard 'WASD'
        printerJogStepSize:tuple[float,float,float] = (0.5, 0.5, 0.25) # x,y,z respectively, in millimeters
        printerPosUpdateInterval = 1/15 # interval between 3d printer position readout (for visualization only)

        positionOffset:tuple[float,float,float] = (116.5, 108.0, 11.25) # IMPORTANT: this is the (relative->absolute) 0-position for this excercise

        DRAW_HIST_LEN = 200 # how many recent datapoints to draw (just for debug). Lower = higher FPS, higher = more points shown

        ############# variables:
        ## commanded positions:
        desiredRelPos:list[float,float,float] = [0.0, 0.0, 0.0] # IMPORTANT: desiredRelPos is RELATIVE TO positionOffset. Absolute position is the sum of both
        ## readout positions (just for visualization (for now)):
        printerTargetPosFeedback:list[float,float,float] = [-1.0,-1.0,-1.0] # to store the output of getCurrentPosition
        printerCurrentPosFeedback:list[float,float,float] = [-1.0,-1.0,-1.0] # to store the output of getCurrentPosition
        global printerIsHomed, steppersDisabled # for keyHandler (interrupt)
        printerIsHomed:bool = False
        steppersDisabled:bool = False

        ## initialize the measurement data array:
        list5D: dict[int,list[tuple[float,float,float,float]]] = {} # {baud : [(x,y,z,data), etc.]}
        for key in baudRatesToTest:
            list5D[key] = [] # init empty array

        ## command-line file loading (mostly because c2Renderer doesn't do file drag-dropping (like pygame does))
        try:
            import sys # used for cmdline arguments
            if(sys.argv[1].endswith(".xlsx") if ((type(sys.argv[1]) is str) if (len(sys.argv) > 1) else False) else False): #a long and convoluted way of checking if a file was (correctly) specified
                print("found sys.argv[1], attempting to import:", sys.argv[1])
                list5D = readFromExcel(sys.argv[1])
                ## there are some things you can do to make the user more confortable
                print("loaded file debug:", len(list5D), list5D.keys())
                # desiredRelPos[2] = max([entry[2] for entry in list5D[tuple(list5D.keys())[-1]]]) # get the highest Z pos used in the test. NOTE: not strictly needed, but it avoid confusion after loading
                # print("set desiredRelPos[2] to:", round(desiredRelPos[2],3), " after loading file")
                DRAW_HIST_LEN *= 2 # increase this, because debugging after-the-fact requires more data than live debugging
                # IRtestItts[2] = len(list5D)-1 # start by looking at the last baud rate from the excel file
                ## this one is just frivolous:
                # import os
                # trimmedFilename = os.path.split(sys.argv[1]) # save the name of the loaded mapfile
                # print("loaded cmdline file:", trimmedFilename)
                ## FINALLY, just open a matplotlib graph by defualt (becuase that's probably the main reason for opening a file)
                plot5D(list5D)
                sys.exit(); # stop the whole code
            elif(len(sys.argv) > 1):
                print("ignored commandline arguments:", sys.argv[1:])
        finally:
            _=0
        
        doMoveMacro:list[bool,float] = [False,-1] # to bet set to True by interrupt functions (indicates that moveMacro() should be run ASAP). NOTE: list to force python to really find it
        printerPosUpdateTimer = time.time()
        
        global IRtestingActive # for keyHandler (interrupt)
        IRtestingActive:bool = False # whether to continue testing
        IRtestItts:list[float,int,int] = [0.0, 0, 0] # (hor_spiral_angle,vert,baud) iterator counters for the IR tests
        
        badDataPattern:list[int] = [0 for i in range(256)] # mostly for debugging

        addPos = lambda posOne, posTwo : [(posOne[i] + posTwo[i]) for i in range(min(len(posOne), len(posTwo)))]
        subtractPos = lambda posOne, posTwo : [(posOne[i] - posTwo[i]) for i in range(min(len(posOne), len(posTwo)))]
        POS_MATCH_THRESH_DEFAULT = 0.04 # (mm) sum of error should not exceed this value for two positions to be considered equal
        matchPos = lambda posOne, posTwo, thresh=POS_MATCH_THRESH_DEFAULT : (sum([abs(entry) for entry in subtractPos(posOne, posTwo)]) <= thresh)

        ## start by printing the COM ports:
        print("serial ports:", [(entry.name, entry.description) for entry in serial.tools.list_ports.comports()])

        ## init printer serial:
        # printerSerial = initSerial("COM7")
        printerSerial = initSerial("COM" + input("please enter the number (only the number) of the COM port of the  3D printer  : COM"), printerBaud)
        if(printerSerial is None):
            print("can't continue without open serial port")
            exit()
        # ## start by waiting for the 3D printer to boot up completely: NOTE: code below is commented out, becuase using .rts=0 and .dtr=0, you can prevent printer reboot on serial init
        # printerInitialized = False; printedData = b''; debugTimer = time.time()
        # print("waiting for 3d printer init phrase")
        # while(not printerInitialized):
        #     newStuff = printerSerial.read_until(b'//action:prompt_end')
        #     if(len(newStuff) > 0):
        #         print(newStuff);  debugTimer = time.time()
        #     printedData += newStuff
        #     if((time.time() - debugTimer) > 1.0):
        #         print("...");  debugTimer = time.time()
        # ## now that the phrase we were looking for has been found, it will still print a few more things we don't care about
        # debugTimer = time.time()
        # while((time.time() - debugTimer) < 1.0):
        #     printedData += printerSerial.read_all()#read(10000000000)
        ## disable auto-report temperatures and position:
        disableAutoReports(printerSerial)
        ## now that the printer is ready to talk:
        # autoHome(printerSerial); time.sleep(10) # TODO: add home checking to autoHome() function
        # goToPos(printerSerial, G0args=(positionOffset, printerSafeFeedrate)) ## you COULD immedietly ask the printer to move to positionOffset... However, i think it's wiser to wait for a human to press ENTER
        print("3d printer init done!")

        ## init IR serial(s):
        IR_serial_timeout = SERIAL_TIMEOUT_DEFAULT + ((3*8)/baudRatesToTest[0]) + (0.020 if (baudRatesToTest[0] < 9600) else 0) # if the baud rate is really low, increasing the timeout a litte might be wise
        # printerSerial = initSerial("COM7") # if you know the COMport beforehand, you could always just 
        IR_RX_serial_port = "COM" + input("please enter the number (only the number) of the COM port of the  IR RX serial  : COM")
        IR_RX_serial = initSerial(IR_RX_serial_port, baudRatesToTest[0], IR_serial_timeout) # NOTE: baud will change in IR tests later (assuming len(baudRatesToTest) > 1)
        IR_TX_serial_port = "COM" + input("please enter the number (only the number) of the COM port of the  IR TX serial  : COM")
        IR_TX_serial = (IR_RX_serial if (IR_TX_serial_port == IR_RX_serial_port) else initSerial(IR_TX_serial_port, baudRatesToTest[0], IR_serial_timeout))
        if(IR_TX_serial_port == IR_RX_serial_port): print("IR RX and TX serial ports are the same! (which is fine, this is just debug)")


        #### functions for handling the movement:
        def moveMacro(feedrate:float=(-1)) -> bool:
            """ just a macro! \n
                writes G0(args) to serialObj and waits for 'ok' respose """
            gcode:bytes = GC.G0(addPos(desiredRelPos, positionOffset), feedrate)
            # print("writing to printer:", gcode)
            printerSerial.write(gcode)
            success, readData = waitForOK(printerSerial)
            if(not success):
                print("moveMacro() unsuccessfull!")
            if(not readData.endswith(GC.GCODE_MARLIN_OK)):
                print("waitForOK() returned:", readData)
            return(success)

        #### functions for IR testing:
        def IRtestUpdateDesiredRelPos(advance:bool=True, CCW:bool=False):
            """ update desiredRelPos. \n
                It moves in a (horizontal) spiral pattern, and will move to the next (vertical) step(/layer) when it looks like no more data can be collected on the current layer \n
                'advance' should only be false if you want to re-affirm/reset desiredPos (without advancing a step in the loop)
                'CCW' just determines the spiral rotation direction. Only needs to be constant/consistant, other than that it shouldn't matter """
            if(advance):
                lastRadius = np.hypot(*desiredRelPos[0:2])
                keepSpiraling = True
                for i in range(len(list5D[baudRatesToTest[IRtestItts[2]]])-1, -1, -1): # scroll through list backwards
                    *xyzPos, measurement = list5D[baudRatesToTest[IRtestItts[2]]][i]
                    if(abs(xyzPos[2] - desiredRelPos[2]) > 0.01): # if the Z position of the previous test is different
                        break # keep spiraling for sure
                    if(measurement > IRtestContinuationThresh): # if (any of) the previous meausrement(s) (within the same vertical step (layer)) contain(s) real data
                        break # keep spiraling untill all measurements are below the threshold for sure
                    if((lastRadius - np.hypot(*xyzPos[0:2])) > IRtestHorizontalStepsize): # if it has been more than 360 degrees (by checking whether the spiral radius has changed 1 full stepsize)
                        ## if no real data has been recorded in 1 full rotation, stop spiraling and move on to the next vertical step (layer)
                        keepSpiraling = False;  break
                if(keepSpiraling):
                    IRtestItts[0] += (IRtestHorizontalStepsize / lastRadius) if (lastRadius > IRtestHorizontalStepsize) else np.deg2rad(60) # constant-arc-length (except for first rotation)
                    if(((IRtestItts[0]/(2*np.pi)) * IRtestHorizontalStepsize) > IRtestHorizontalDistMax): # if next radius would exceed manually set limit (unlikely)
                        print("(debug): IRtestHorizontalDistMax reached")
                        keepSpiraling = False
                if(not keepSpiraling): # move to the next vertical step (layer)
                    IRtestItts[0] = 0.0 # reset angle to 0 (radians)
                    IRtestItts[1] += 1 # vertical step
                    if((IRtestItts[1] * IRtestVerticalStepsize) > IRtestVerticalDistMax): # if the next vertical position is above the maximum
                        return(True) # the whole range of motion has been completed
                    for i in range(len(list5D[baudRatesToTest[IRtestItts[2]]])-1, -1, -1): # scroll through list backwards
                        *xyzPos, measurement = list5D[baudRatesToTest[IRtestItts[2]]][i]
                        if(measurement > IRtestContinuationThresh): # if (any of) the previous meausrement(s) contain(s) real data
                            break
                        if((desiredRelPos[2] - xyzPos[2]) >= IRtestVertStopThresh): # the datapoint in question is this much lower that the current one
                            # if no measurements have been recorded in the last several vertical steps (layers), consider the test concluded (there's hardly any point in doing more measurements)
                            return(True) # it is unlikely that any good data will be recorded at this point
            desiredRelPos[0] = (-1 if CCW else 1) * np.sin(IRtestItts[0]) * ((IRtestItts[0]/(2*np.pi)) * IRtestHorizontalStepsize) # spiral (inspired by my PCBcoilV2.circularSpiral.calcPos())
            desiredRelPos[1] =          1         * np.cos(IRtestItts[0]) * ((IRtestItts[0]/(2*np.pi)) * IRtestHorizontalStepsize) # spiral
            desiredRelPos[2] = IRtestItts[1] * IRtestVerticalStepsize
            return(False) # returns whether the whole range of motion has been completed (if it reached this point, then it hasn't)
        
        ##### visualization stuff:
        import cv2Renderer as rend
        ## some UI window initialization
        windowHandler = rend.cv2WindowHandler([1280, 720], "IR_alignment_gcode")
        drawer = rend.cv2Drawer(windowHandler, sizeScale=100) # only 1 renderer in the window
        def keyHandler(keycode:int, drawer:rend.cv2Drawer):
            """ handles key presses """
            ## NOTE: some keys are used by cv2Drawer class by default: 'z'=zoom, 'g'=grid
            char = chr(keycode) # just interprets an ascii table, basically
            global IRtestingActive
            if(char == 'r'): # r -> reset to zero-relative-pos
                if(not IRtestingActive): 
                    desiredRelPos[0]=0;desiredRelPos[1]=0;desiredRelPos[2]=0 # reset relative position
                    doMoveMacro[0] = True; doMoveMacro[1] = printerSafeFeedrate
            elif(char == 'h'): # h -> auto-home
                if(not IRtestingActive):
                    global printerIsHomed
                    printerIsHomed = autoHome(printerSerial)
            elif(char == 'l'): # l(L) -> disable steppers
                if(not IRtestingActive):
                    global steppersDisabled
                    steppersDisabled = disableSteppers(printerSerial)
            elif(char == 'w'): # w -> forwards
                if(not IRtestingActive):  desiredRelPos[1] += printerJogStepSize[1];  doMoveMacro[0] = True;  doMoveMacro[1] = printerJogFeedrate
            elif(char == 'a'): # a -> left
                if(not IRtestingActive):  desiredRelPos[0] -= printerJogStepSize[0];  doMoveMacro[0] = True;  doMoveMacro[1] = printerJogFeedrate
            elif(char == 's'): # s -> backwards
                if(not IRtestingActive):  desiredRelPos[1] -= printerJogStepSize[1];  doMoveMacro[0] = True;  doMoveMacro[1] = printerJogFeedrate
            elif(char == 'd'): # d -> right
                if(not IRtestingActive):  desiredRelPos[0] += printerJogStepSize[0];  doMoveMacro[0] = True;  doMoveMacro[1] = printerJogFeedrate
            elif(char == 'q'): # q -> down
                if(not IRtestingActive):  desiredRelPos[2] -= printerJogStepSize[2];  doMoveMacro[0] = True;  doMoveMacro[1] = printerJogFeedrate
            elif(char == 'e'): # e -> up
                if(not IRtestingActive):  desiredRelPos[2] += printerJogStepSize[2];  doMoveMacro[0] = True;  doMoveMacro[1] = printerJogFeedrate
            elif(char == ' '): # SPACE
                IRtestingActive = not IRtestingActive # pause/unpause testing
                if(IRtestingActive):
                    IRtestUpdateDesiredRelPos( advance=False ) # should reset the desiredPos to the last point (without actually advancing)
                    doMoveMacro[0] = True; doMoveMacro[1] = printerSafeFeedrate
            elif(char == 'v'): # v -> (view) graph
                if(not IRtestingActive): # just to avoid stalling the test
                    plot5D(list5D)
            elif(char == 'k'): # k -> save to excel (only meant for interrupted tests) 
                # if(not IRtestingActive): # still not recommended to do while test is active, due to the interrupting nature (you gotta get into semaphores for that)
                saveToExcel(list5D, generateFileName(list5D))
            elif((char != 'z') and (char != 'g')): # 'z' and 'g' are (currently) used by the cv2Drawer (which preceeds this function)
                print("unused keycode:", keycode, char)
        drawer.keyboardCallbackFunc = keyHandler # whenever a key is pressed, cv2 will catch it and call the keyHander() function (after calling 2 other functions from the classes, btw)
        # targetFPS = 60 # manually enforced at the end of every loop, to limit CPU usage
        
        ## visualization loop:
        while(windowHandler.keepRunning):
            loopStart = time.time()
            
            if(doMoveMacro[0]): # for calling printer commands from interrupt functions (without interference)
                doMoveMacro[0] = False; doMoveMacro[1] = -1
                moveMacro(doMoveMacro[1])

            # update printer position feedback data (but not every frame, that would be excessive)
            if((loopStart - printerPosUpdateTimer) > printerPosUpdateInterval):
                printerPosUpdateTimer = loopStart
                success, printerTargetPosFeedback, printerCurrentPosFeedback = getCurrentPosition(printerSerial)

            if(IRtestingActive): ## the actual testing loop
                ## start by doing a measurement at the current position
                if(matchPos(desiredRelPos, subtractPos(printerCurrentPosFeedback, positionOffset))):
                    measurement = np.average([IRresponseTest(IR_TX_serial, IR_RX_serial) for _ in range(IRtestPasses)]) # perform the actual test
                    list5D[baudRatesToTest[IRtestItts[2]]].append((*desiredRelPos,measurement))
                    print("measurement:", stringifyPos(list5D[baudRatesToTest[IRtestItts[2]]][-1][0:3]), round(measurement,3), int(measurement*256))
                    switchToNextBaud = IRtestUpdateDesiredRelPos() # updated desiredRelPos
                    if(abs(desiredRelPos[2] - printerCurrentPosFeedback[2]) > 0.01): # if it's about to move vertically
                        temp = desiredRelPos[0:2]; desiredRelPos[0]=(printerTargetPosFeedback[0]-positionOffset[0]); desiredRelPos[1]=(printerTargetPosFeedback[1]-positionOffset[1]) # use current x,y position
                        moveMacro() # insert an extra move, which should move exclusively upwards # which the printer likes a little better
                        desiredRelPos[0]=temp[0]; desiredRelPos[1]=temp[1] # now restore the calculated position (which should just be (0,0), but still)
                    moveMacro()
                    if(switchToNextBaud):
                        IRtestItts[2] += 1
                        if(IRtestItts[2] >= len(baudRatesToTest)):
                            IRtestItts[2] -= 1 # (just to avoid index overflow in drawing code)
                            print("testing done!")
                            IRtestingActive = False
                            try:
                                saveToExcel(list5D, generateFileName(list5D))
                            except Exception as excep:
                                print("failed to save to excel!", excep)
                            try:
                                plot5D(list5D)
                            except Exception as excep:
                                print("failed to plot in matplotlib!:", excep)
                            # continue
                        else: # if there are more baud rates to test
                            # time.sleep(0.1) # wait an extra 100ms before changing baud, to let the UART IC send any last data still in the buffer (commented out, as IRresponseTest() reads all data)
                            IR_RX_serial.baudrate = baudRatesToTest[IRtestItts[2]] # will call _reconfigure_port() underwater (may result in unintended pulse, and therefore some garbage data)
                            # if(IR_TX_serial_port != IR_RX_serial_port): # extra check is nice, but not strictly needed
                            IR_TX_serial.baudrate = baudRatesToTest[IRtestItts[2]]
                            if(IR_RX_serial.timeout > SERIAL_TIMEOUT_DEFAULT):
                                IR_serial_timeout = SERIAL_TIMEOUT_DEFAULT + (0.015 if (baudRatesToTest[0] < 9600) else 0) #also update timeout (in case you can go faster as a result)
                                IR_RX_serial.timeout = IR_serial_timeout
                            time.sleep(0.1) # wait 100ms, just for good measure
                            IR_RX_serial.flush()
                            # while(IR_RX_serial.in_waiting > 0):     IR_RX_serial.read() # manual flush

            drawer.background() # draw background
            
            ## draw the observed data as small dots, just to get a preview of what it might look like when its done
            stopIndex = len([None for entry in list5D[baudRatesToTest[IRtestItts[2]]] if (desiredRelPos[2] >= entry[2])]) # find the highest index where the Z position is below/at the current desired Z pos
            for i in range(max(stopIndex-DRAW_HIST_LEN, 0), stopIndex, 1): # scroll through list from older to newest
                *xyzPos, measurement = list5D[baudRatesToTest[IRtestItts[2]]][i]
                color = [  0,int(min(255,measurement*512)),int(min(255,512-(measurement*512)))] # [B,R,G] transitions red->yellow->green based on measurement 0.0->1.0
                radius = 0.1 + (0.05 * (desiredRelPos[2] - xyzPos[2]) / IRtestVerticalStepsize) # the more Z distance to the measurement, the bigger the circle
                if((radius <= 0.05) or (radius >= 0.3)): continue #radius = 0.1   # very niche fix, only applies if you manually jog the head AFTER recording data above that coordinate
                drawer.drawCircle(xyzPos[0:2], radius, color) # draw datapoint

            ## draw the current positions as clearly visible dots
            drawer.drawCircle(subtractPos(printerCurrentPosFeedback[0:2], positionOffset), 0.5, [255,  0,  0]) # draw feedback current position (blue)
            drawer.drawCircle(subtractPos(printerTargetPosFeedback[0:2], positionOffset) , 0.4, [  0,255,255]) # draw feedback target position (yellow)
            drawer.drawCircle(desiredRelPos[0:2]                                         , 0.3, [255,  0,255]) # draw current desired position (purple)

            ## text on screen
            stringifyPos = lambda pos, decimals=2 : str([round(entry,decimals) for entry in pos])
            drawer.statStrings = [] # reset text in topleft corner
            drawer.statStrings.append(stringifyPos(desiredRelPos)) # show desired pos
            drawer.statStrings.append(stringifyPos(printerTargetPosFeedback) + "=" + stringifyPos(subtractPos(printerTargetPosFeedback, positionOffset))) # show target pos (feedback)
            drawer.statStrings.append(stringifyPos(printerCurrentPosFeedback) + "=" + stringifyPos(subtractPos(printerCurrentPosFeedback, positionOffset))) # show current pos (feedback)
            drawer.statStrings.append("homed: "+str(printerIsHomed)) # show current pos (feedback)
            drawer.statStrings.append("steppersDisabled: "+str(steppersDisabled)) # show current pos (feedback)
            drawer.statStrings.append("IRtestingActive: "+str(IRtestingActive)) # just debug
            if(IRtestingActive):
                drawer.statStrings.append("[" + str(round(np.rad2deg(IRtestItts[0]),1)) + "," + str(IRtestItts[1]) + "," + str(IRtestItts[2]) + "]") # debug IRtestItts (to show progress)
                progressPercentage = (IRtestVerticalStepsize / IRtestVerticalDistMax) * ((((IRtestItts[0]/(2*np.pi)) * IRtestHorizontalStepsize) / IRtestHorizontalDistMax) + IRtestItts[1]) # approximate completion
                drawer.statStrings.append("progess:~" + str(round(progressPercentage*100)) + "%")

            drawer.renderFG() # draw foreground (text and stuff)
            # drawer.redraw() # render all elements
            windowHandler.frameRefresh()
            
            # loopEnd = time.time() #this is only for the 'framerate' limiter (time.sleep() doesn't accept negative numbers, this solves that)
            # if((loopEnd-loopStart) < (1/(targetFPS*1.05))): #FPS limiter (optional)
            #     time.sleep((1/targetFPS)-(loopEnd-loopStart))
            # # elif((loopEnd-loopStart) > (1/5)):
            # #     print("main process running slow", 1/(loopEnd-loopStart))
    finally:
        try:
            windowHandler.end() # correctly shut down cv2 window
            print("drawer stopping done")
        except Exception as excep:
            print("couldn't run cv2 window end():", excep)
        try:
            printerSerial.close()
            print("closed printerSerial")
        except Exception as excep:
            print("couldn't close printerSerial", excep)
        try:
            IR_RX_serial.close()
            print("closed IR_RX_serial")
        except Exception as excep:
            print("couldn't close IR_RX_serial", excep)
        try:
            IR_TX_serial.close()
            print("closed IR_TX_serial")
        except Exception as excep:
            print("couldn't close IR_RX_serial", excep)
        try:
            saveToExcel(list5D)
            print("saved to excel file:", saveToExcel.__defaults__[0])
        except Exception as excep:
            print("couldn't save data to excel", excep)